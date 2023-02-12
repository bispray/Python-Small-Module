# coding: utf-8

from openpyxl import load_workbook
import json
import requests
import time
import os
import sys
from configparser import ConfigParser

def translate():
    cp = ConfigParser()
    cp.read('./translateConfig.cfg')
    newVocabularies = cp.get('file_dir', 'newVocabularies')
    print('\033[32m要装载的文件位置-> \033[0m'+ newVocabularies)
    start_time = time.time()

    wb = load_workbook(newVocabularies)
    table = wb['Sheet1']
    rows = table.max_row

    headers = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) ""AppleWebKit/537.36 "}
    url = "https://fanyi.baidu.com/sug"
    testr = requests.get("https://fanyi.baidu.com", headers=headers)
    if testr.status_code == 200:
        print('\033[32m[+]\033[0m' + ' Communication is normal, the next operation is about to start')

    for i in range(1, rows+1):
        Data = table.cell(i, 1).value

        data = {"kw":Data}
        r = requests.post(url = url, headers = headers, data = data)
        jiexi = json.loads(r.text)

        table.cell(i,2).value = jiexi['data'][0]['v']

        spend_time = time.time() - start_time
        if i != rows:
            print('已完成装载' + str(i) + '个，剩余' + str(rows - i) + '个，总计用时' + str(spend_time), end='\r')
        elif i == rows:
            print('已完成装载' + str(i) + '个，剩余0个，总计用时' + str(spend_time), end='')

        time.sleep(0.2)

    wb.save(newVocabularies)

if __name__ == '__main__':
    translate()